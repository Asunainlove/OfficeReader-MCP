"""
MCP Server for OfficeReader - Convert Office documents to Markdown.
Supports: Word (docx/doc), Excel (xlsx/xls), PowerPoint (pptx/ppt)
"""

import asyncio
import json
import os
import sys
from pathlib import Path
from typing import Any

from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import (
    TextContent,
    Tool,
    INVALID_PARAMS,
    INTERNAL_ERROR,
)

from .converter import OfficeConverter, SUPPORTED_FORMATS
from .config_loader import get_config

# Initialize the MCP server
server = Server("officereader-mcp")

# Global converter instance
_converter: OfficeConverter | None = None


def get_converter() -> OfficeConverter:
    """Get or create the converter instance."""
    global _converter
    if _converter is None:
        # Load configuration
        config = get_config()
        cache_dir = config.get_cache_dir()
        _converter = OfficeConverter(cache_dir=cache_dir)
    return _converter


def get_supported_extensions() -> list[str]:
    """Get list of all supported file extensions."""
    extensions = []
    for exts in SUPPORTED_FORMATS.values():
        extensions.extend(exts)
    return extensions


@server.list_tools()
async def list_tools() -> list[Tool]:
    """List available tools for the MCP server."""
    supported_exts = ", ".join(get_supported_extensions())

    return [
        Tool(
            name="convert_document",
            description=f"""Convert Office documents to Markdown format for Claude to read.

Supported formats:
- Word: .docx, .doc
- Excel: .xlsx, .xls (converts each sheet to a Markdown table)
- PowerPoint: .pptx, .ppt (extracts text and images from slides)

Features:
- Text extraction with formatting (headings, bold, italic, lists, tables)
- Image extraction and optimization (auto-compressed for efficiency)
- Speaker notes extraction from PowerPoint
- Multi-sheet support for Excel

Images are automatically optimized (WebP format, max 1920x1080) to reduce size
while maintaining readability for Claude.""",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": f"Absolute path to the Office document. Supported: {supported_exts}",
                    },
                    "extract_images": {
                        "type": "boolean",
                        "description": "Whether to extract and include images (default: true)",
                        "default": True,
                    },
                    "image_format": {
                        "type": "string",
                        "enum": ["file", "base64", "both"],
                        "description": "How to handle images: 'file' saves to disk (recommended), 'base64' embeds in markdown, 'both' does both",
                        "default": "file",
                    },
                    "output_name": {
                        "type": "string",
                        "description": "Custom name for the output (without extension). If not provided, generates from filename.",
                    },
                },
                "required": ["file_path"],
            },
        ),
        Tool(
            name="read_converted_markdown",
            description="""Read the content of a previously converted markdown file.

Use this after convert_document to get the actual markdown content.
This is useful when you want to process or analyze the converted document.""",
            inputSchema={
                "type": "object",
                "properties": {
                    "markdown_path": {
                        "type": "string",
                        "description": "Path to the markdown file (returned by convert_document)",
                    },
                },
                "required": ["markdown_path"],
            },
        ),
        Tool(
            name="list_conversions",
            description="""List all cached document conversions.

Shows all documents that have been converted, including their output paths
and number of extracted images. Useful for finding previously converted files.""",
            inputSchema={
                "type": "object",
                "properties": {},
            },
        ),
        Tool(
            name="clear_cache",
            description="""Clear all cached conversions.

Removes all converted markdown files and extracted images from the cache.
Use this to free up disk space or reset the conversion cache.""",
            inputSchema={
                "type": "object",
                "properties": {},
            },
        ),
        Tool(
            name="get_document_metadata",
            description=f"""Get metadata from an Office document without full conversion.

Extracts document properties like title, author, creation date, etc.
Faster than full conversion when you only need metadata.

Supported formats: {supported_exts}""",
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": f"Absolute path to the Office document. Supported: {supported_exts}",
                    },
                },
                "required": ["file_path"],
            },
        ),
        Tool(
            name="get_supported_formats",
            description="""Get list of all supported file formats.

Returns a dictionary of file types (word, excel, powerpoint) and their extensions.""",
            inputSchema={
                "type": "object",
                "properties": {},
            },
        ),
    ]


def get_cache_location_notice(converter: OfficeConverter) -> str:
    """Generate a notice about cache location."""
    return f"[Cache Location] {converter.cache_dir}"


@server.call_tool()
async def call_tool(name: str, arguments: dict[str, Any]) -> list[TextContent]:
    """Handle tool calls."""
    try:
        converter = get_converter()
        cache_notice = get_cache_location_notice(converter)

        if name == "convert_document":
            file_path = arguments.get("file_path")
            if not file_path:
                return [TextContent(
                    type="text",
                    text=json.dumps({"error": "file_path is required"}, ensure_ascii=False)
                )]

            extract_images = arguments.get("extract_images", True)
            image_format = arguments.get("image_format", "file")
            output_name = arguments.get("output_name")

            result = converter.convert_file(
                file_path=file_path,
                extract_images=extract_images,
                image_format=image_format,
                output_name=output_name,
            )

            # Return summary with markdown preview
            response = {
                "status": "success",
                "file_type": result.get("file_type", "unknown"),
                "cache_location": str(converter.cache_dir),
                "output_path": result["output_path"],
                "conversion_dir": result["conversion_dir"],
                "images_extracted": len(result["images"]),
                "image_paths": result["images"],
                "metadata": result["metadata"],
                "markdown_preview": result["markdown"][:2000] + "..." if len(result["markdown"]) > 2000 else result["markdown"],
                "markdown_length": len(result["markdown"]),
                "hint": "Use 'list_conversions' to see all cached files, or 'read_converted_markdown' to read the full content.",
            }

            return [TextContent(
                type="text",
                text=f"{cache_notice}\n\n" + json.dumps(response, ensure_ascii=False, indent=2)
            )]

        elif name == "read_converted_markdown":
            markdown_path = arguments.get("markdown_path")
            if not markdown_path:
                return [TextContent(
                    type="text",
                    text=f"{cache_notice}\n\n" + json.dumps({"error": "markdown_path is required"}, ensure_ascii=False)
                )]

            path = Path(markdown_path)
            if not path.exists():
                return [TextContent(
                    type="text",
                    text=f"{cache_notice}\n\n" + json.dumps({"error": f"File not found: {markdown_path}"}, ensure_ascii=False)
                )]

            with open(path, "r", encoding="utf-8") as f:
                content = f.read()

            return [TextContent(
                type="text",
                text=f"{cache_notice}\n[Reading] {markdown_path}\n[Length] {len(content)} characters\n\n---\n\n{content}"
            )]

        elif name == "list_conversions":
            cache_info = converter.get_cache_info()

            # Build a formatted summary
            summary_lines = [
                "=" * 50,
                "       OfficeReader-MCP Cache Information",
                "=" * 50,
                "",
                cache_notice,
                f"Output directory: {cache_info.get('output_dir', 'N/A')}",
                "",
                f"Total cached conversions: {cache_info['total_conversions']}",
                f"Total cache size: {cache_info.get('total_size_human', 'N/A')}",
                "",
            ]

            if cache_info['conversions']:
                summary_lines.append("-" * 50)
                summary_lines.append("       Cached Documents")
                summary_lines.append("-" * 50)

                for i, conv in enumerate(cache_info['conversions'], 1):
                    summary_lines.append(f"\n[{i}] {conv['name']}")
                    summary_lines.append(f"    Directory: {conv['path']}")
                    if conv['markdown_files']:
                        summary_lines.append(f"    Markdown:  {conv['markdown_files'][0]}")
                    summary_lines.append(f"    Images:    {conv['image_count']} files")
                    summary_lines.append(f"    Size:      {conv.get('size_human', 'N/A')}")
                    if conv.get('modified'):
                        summary_lines.append(f"    Modified:  {conv['modified']}")
            else:
                summary_lines.append("-" * 50)
                summary_lines.append("No cached conversions found.")
                summary_lines.append("")
                summary_lines.append("To convert a document, use the 'convert_document' tool:")
                summary_lines.append("  file_path: <path to your .docx or .doc file>")

            summary_lines.append("")
            summary_lines.append("=" * 50)
            summary_lines.append("\n--- Raw JSON Data ---")

            return [TextContent(
                type="text",
                text="\n".join(summary_lines) + "\n" + json.dumps(cache_info, ensure_ascii=False, indent=2)
            )]

        elif name == "clear_cache":
            result = converter.clear_cache()
            response = {
                "status": "success",
                "cache_location": str(converter.cache_dir),
                "cleared_count": result["count"],
                "cleared_directories": result["cleared"],
            }
            return [TextContent(
                type="text",
                text=f"{cache_notice}\n\nCache cleared successfully!\n\n" + json.dumps(response, ensure_ascii=False, indent=2)
            )]

        elif name == "get_document_metadata":
            file_path = arguments.get("file_path")
            if not file_path:
                return [TextContent(
                    type="text",
                    text=f"{cache_notice}\n\n" + json.dumps({"error": "file_path is required"}, ensure_ascii=False)
                )]

            from .converter import get_file_type

            file_path_obj = Path(file_path)
            file_type = get_file_type(file_path_obj)

            metadata = {
                "file": file_path,
                "file_type": file_type,
                "cache_location": str(converter.cache_dir),
            }

            if file_type == "word":
                from docx import Document
                doc = Document(file_path)
                core_props = doc.core_properties
                metadata.update({
                    "title": core_props.title or "",
                    "author": core_props.author or "",
                    "created": str(core_props.created) if core_props.created else "",
                    "modified": str(core_props.modified) if core_props.modified else "",
                    "last_modified_by": core_props.last_modified_by or "",
                    "subject": core_props.subject or "",
                    "keywords": core_props.keywords or "",
                    "category": core_props.category or "",
                    "comments": core_props.comments or "",
                    "revision": core_props.revision,
                })
            elif file_type == "excel":
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                props = wb.properties
                metadata.update({
                    "title": props.title or "",
                    "creator": props.creator or "",
                    "created": str(props.created) if props.created else "",
                    "modified": str(props.modified) if props.modified else "",
                    "sheet_count": len(wb.sheetnames),
                    "sheet_names": wb.sheetnames,
                })
            elif file_type == "powerpoint":
                from pptx import Presentation
                prs = Presentation(file_path)
                core_props = prs.core_properties
                metadata.update({
                    "title": core_props.title or "",
                    "author": core_props.author or "",
                    "created": str(core_props.created) if core_props.created else "",
                    "modified": str(core_props.modified) if core_props.modified else "",
                    "subject": core_props.subject or "",
                    "slide_count": len(prs.slides),
                })
            else:
                return [TextContent(
                    type="text",
                    text=f"{cache_notice}\n\n" + json.dumps({
                        "error": f"Unsupported file format: {file_path_obj.suffix}",
                        "supported": get_supported_extensions()
                    }, ensure_ascii=False)
                )]

            return [TextContent(
                type="text",
                text=f"{cache_notice}\n\n" + json.dumps(metadata, ensure_ascii=False, indent=2)
            )]

        elif name == "get_supported_formats":
            return [TextContent(
                type="text",
                text=json.dumps({
                    "supported_formats": SUPPORTED_FORMATS,
                    "all_extensions": get_supported_extensions(),
                }, ensure_ascii=False, indent=2)
            )]

        else:
            return [TextContent(
                type="text",
                text=f"{cache_notice}\n\n" + json.dumps({"error": f"Unknown tool: {name}"}, ensure_ascii=False)
            )]

    except FileNotFoundError as e:
        converter = get_converter()
        return [TextContent(
            type="text",
            text=f"[Cache Location] {converter.cache_dir}\n\n" + json.dumps({"error": str(e)}, ensure_ascii=False)
        )]
    except ValueError as e:
        converter = get_converter()
        return [TextContent(
            type="text",
            text=f"[Cache Location] {converter.cache_dir}\n\n" + json.dumps({"error": str(e)}, ensure_ascii=False)
        )]
    except Exception as e:
        converter = get_converter()
        return [TextContent(
            type="text",
            text=f"[Cache Location] {converter.cache_dir}\n\n" + json.dumps({
                "error": f"Conversion failed: {str(e)}",
                "error_type": type(e).__name__,
            }, ensure_ascii=False)
        )]


def main():
    """Main entry point for the MCP server."""
    async def run():
        async with stdio_server() as (read_stream, write_stream):
            await server.run(
                read_stream,
                write_stream,
                server.create_initialization_options(),
            )

    asyncio.run(run())


if __name__ == "__main__":
    main()
