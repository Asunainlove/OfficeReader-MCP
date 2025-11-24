"""
Excel to Markdown converter with image extraction.
"""

import base64
from datetime import datetime
from pathlib import Path
from typing import Optional
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

from .image_optimizer import ImageOptimizer


class ExcelConverter:
    """Convert Excel files (.xlsx, .xls) to Markdown format."""

    def __init__(self, cache_dir: Path, image_optimizer: ImageOptimizer):
        """
        Initialize the Excel converter.

        Args:
            cache_dir: Directory for caching outputs
            image_optimizer: Image optimization utility
        """
        self.cache_dir = cache_dir
        self.image_optimizer = image_optimizer

    def convert_file(
        self,
        file_path: str,
        extract_images: bool = True,
        image_format: str = "file",
        output_name: Optional[str] = None,
        conversion_dir: Path = None,
        images_dir: Path = None,
    ) -> dict:
        """
        Convert Excel file to Markdown.

        Args:
            file_path: Path to Excel file
            extract_images: Whether to extract embedded images
            image_format: "file", "base64", or "both"
            output_name: Custom output name
            conversion_dir: Directory for this conversion
            images_dir: Directory for extracted images

        Returns:
            dict with markdown content, images, and metadata
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()
        if suffix not in [".xlsx", ".xls"]:
            raise ValueError(f"Unsupported file format: {suffix}")

        # Load workbook
        wb = load_workbook(file_path, data_only=True)

        markdown_parts = []
        all_images = []
        image_counter = 0

        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Add sheet header
            markdown_parts.append(f"# Sheet: {sheet_name}\n")

            # Extract data from sheet
            data = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None and str(cell).strip() for cell in row):
                    data.append(row)

            if data:
                # Convert to Markdown table
                md_table = self._create_markdown_table(data)
                markdown_parts.append(md_table)
            else:
                markdown_parts.append("*Empty sheet*\n")

            # Extract images from sheet
            if extract_images and hasattr(ws, "_images"):
                for img in ws._images:
                    image_counter += 1
                    img_name = f"excel_image_{image_counter:03d}"

                    # Get image data
                    img_bytes = img._data()
                    pil_img = Image.open(BytesIO(img_bytes))

                    # Optimize image
                    optimized_data, ext = self.image_optimizer.optimize_image(
                        pil_img, img_name
                    )

                    img_filename = f"{img_name}{ext}"
                    img_path = images_dir / img_filename

                    if image_format in ["file", "both"]:
                        with open(img_path, "wb") as f:
                            f.write(optimized_data)
                        all_images.append(str(img_path))
                        markdown_parts.append(f"\n![{img_name}](images/{img_filename})\n")

                    if image_format in ["base64", "both"]:
                        b64_data = base64.b64encode(optimized_data).decode("utf-8")
                        content_type = f"image/{ext.lstrip('.')}"
                        if image_format == "base64":
                            markdown_parts.append(
                                f"\n![{img_name}](data:{content_type};base64,{b64_data})\n"
                            )

            markdown_parts.append("\n---\n")

        # Extract metadata
        metadata = self._extract_metadata(wb, file_path)

        markdown = "\n".join(markdown_parts)

        return {
            "markdown": markdown,
            "images": all_images,
            "metadata": metadata,
        }

    def _create_markdown_table(self, data: list) -> str:
        """
        Convert 2D data array to Markdown table.

        Args:
            data: List of row tuples

        Returns:
            Markdown formatted table
        """
        if not data:
            return ""

        # Find maximum column count
        max_cols = max(len(row) for row in data)

        # Normalize rows to same length
        normalized_data = []
        for row in data:
            normalized_row = list(row) + [None] * (max_cols - len(row))
            normalized_data.append(normalized_row)

        md_rows = []

        for i, row in enumerate(normalized_data):
            cells = []
            for cell in row:
                # Convert cell value to string
                cell_text = str(cell) if cell is not None else ""
                # Escape pipe characters
                cell_text = cell_text.replace("|", "\\|").replace("\n", " ")
                cells.append(cell_text)

            md_rows.append("| " + " | ".join(cells) + " |")

            # Add header separator after first row
            if i == 0:
                separator = "| " + " | ".join(["---"] * len(cells)) + " |"
                md_rows.append(separator)

        return "\n".join(md_rows) + "\n"

    def _extract_metadata(self, workbook, file_path: Path) -> dict:
        """Extract Excel file metadata."""
        props = workbook.properties

        return {
            "source": str(file_path),
            "converted_at": datetime.now().isoformat(),
            "title": props.title or "",
            "creator": props.creator or "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
        }
